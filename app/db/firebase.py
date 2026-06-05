import firebase_admin
from firebase_admin import credentials, db
from dotenv import load_dotenv
import os
import json
import base64  # Importar base64
import binascii
from collections.abc import Mapping, Sequence
from typing import Any, Dict, List, Optional
from datetime import datetime, timedelta, timezone
from urllib.parse import unquote
import io
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

if os.getenv("VERCEL") != "1":
    print("Cargando variables de entorno desde .env (Modo Local)...")
    load_dotenv(override=os.getenv("SKIP_FIREBASE_INIT", "false").lower() not in {"1", "true", "yes"})
else:
    print("Saltando load_dotenv() (Modo Vercel)...")

# ======================================================================
# INICIALIZACIÓN DE FIREBASE (Modificado para Base64)
# ======================================================================

skip_firebase_init = os.getenv("SKIP_FIREBASE_INIT", "false").lower() in {"1", "true", "yes"}
database_url = os.getenv("FIREBASE_DATABASE_URL")
cred_path = os.getenv("FIREBASE_CREDENTIALS_PATH")

if not skip_firebase_init and not database_url:
    raise ValueError("ERROR FATAL: FIREBASE_DATABASE_URL no está configurada en el entorno.")


def _load_service_account_from_env() -> Optional[dict[str, Any]]:
    cred_json_base64 = os.getenv("FIREBASE_PRIVATE_KEY_JSON_BASE64")
    cred_json_raw = os.getenv("FIREBASE_PRIVATE_KEY_JSON")

    if cred_json_base64:
        decoded_json_string = base64.b64decode(cred_json_base64).decode("utf-8")
        return json.loads(decoded_json_string)

    if cred_json_raw:
        try:
            decoded_json_string = base64.b64decode(cred_json_raw, validate=True).decode("utf-8")
            return json.loads(decoded_json_string)
        except (binascii.Error, UnicodeDecodeError, json.JSONDecodeError):
            return json.loads(cred_json_raw)

    return None

if skip_firebase_init:
    print("Saltando inicialización de Firebase por configuración de entorno.")
elif not firebase_admin._apps:
    try:
        cred = None
        service_account_info = _load_service_account_from_env()
        if service_account_info:
            print("Inicializando Firebase con credenciales JSON desde variable de entorno...")
            cred = credentials.Certificate(service_account_info)
            
        elif cred_path:
            print(f"Inicializando Firebase con ruta de archivo: {cred_path} (Modo Local)...")
            if not os.path.exists(cred_path):
                raise FileNotFoundError(f"El archivo de credenciales no se encuentra en la ruta: {cred_path}")
            cred = credentials.Certificate(cred_path)
        else:
            raise ValueError("No se encontró 'FIREBASE_PRIVATE_KEY_JSON_BASE64', 'FIREBASE_PRIVATE_KEY_JSON' ni 'FIREBASE_CREDENTIALS_PATH'. Revisa tu .env")

        firebase_admin.initialize_app(cred, {
            'databaseURL': database_url
        })
        print("Firebase initialized successfully")
        
    except Exception as e:
        # Imprime un error más detallado si falla la decodificación o inicialización
        print(f"ERROR FATAL AL INICIALIZAR FIREBASE: {str(e)}")
        raise Exception(f"Failed to initialize Firebase: {str(e)}")

# ======================================================================
# ¡FUNCIÓN DE DETECCIÓN MODIFICADA!
# ======================================================================

def _read_float_env(name: str, default: float) -> float:
    raw_value = os.getenv(name, str(default))
    try:
        return float(raw_value)
    except ValueError:
        return default


BRANCH_RESIDUAL_MAX_CURRENT_A = _read_float_env("BRANCH_RESIDUAL_MAX_CURRENT_A", 0.16)
BRANCH_ONE_PC_MAX_CURRENT_A = _read_float_env("BRANCH_ONE_PC_MAX_CURRENT_A", 0.20)
BRANCH_ONE_PC_WORKLOAD_MAX_CURRENT_A = _read_float_env("BRANCH_ONE_PC_WORKLOAD_MAX_CURRENT_A", 0.23)
BRANCH_TWO_PC_MAX_CURRENT_A = _read_float_env("BRANCH_TWO_PC_MAX_CURRENT_A", 0.30)


def _legacy_detect_device_type(irms: float, threshold: float) -> dict:
    """
    Detecta el tipo de dispositivo basándose en el consumo Y EL UMBRAL.
    Retorna: {type, icon, description, color}
    """
    
    # 1. ¡REVISAR SOBRECARGA PRIMERO!
    if irms >= threshold:
        # Sobrecarga masiva (posible cortocircuito)
        if irms >= 15.0: 
            return {
                "type": "¡PICO EXTREMO!",
                "icon": "💥",
                "description": f"Cortocircuito o falla grave detectada ({irms:.2f}A)",
                "color": "#ff0000"
            }
        # Sobrecarga "normal"
        else:
            return {
                "type": "SOBRECARGA",
                "icon": "⚠️",
                "description": f"Consumo ({irms:.2f}A) supera el umbral ({threshold:.1f}A)",
                "color": "#e74c3c" # Rojo peligro
            }

    # 2. SI NO ES SOBRECARGA, identificar el dispositivo
    if irms < 0.01:
        return {
            "type": "Sin carga",
            "icon": "🔌",
            "description": "No hay dispositivos conectados",
            "color": "#95a5a6"
        }
    elif 0.01 <= irms < 0.1:
        return {
            "type": "Audífonos / Carga baja",
            "icon": "🎧",
            "description": "Carga de audífonos o dispositivo de bajo consumo",
            "color": "#3498db"
        }
    elif 0.1 <= irms < 1.5:
        return {
            "type": "Cargador de celular",
            "icon": "📱",
            "description": "Smartphone o tablet en carga",
            "color": "#27ae60"
        }
    elif 1.5 <= irms < 4.0:
        return {
            "type": "Laptop",
            "icon": "💻",
            "description": "Laptop en uso o carga",
            "color": "#f39c12"
        }
    elif 4.0 <= irms < 8.0:
        return {
            "type": "PC de escritorio",
            "icon": "🖥️",
            "description": "Computadora de escritorio (CPU + Monitor)",
            "color": "#e67e22"
        }
    
    # 3. Rango entre "PC" y el umbral: Carga alta pero segura
    elif 8.0 <= irms < threshold:
        return {
            "type": "Múltiples dispositivos",
            "icon": "⚡",
            "description": "Varios dispositivos conectados o carga alta",
            "color": "#e67e22" # Naranja (advertencia, no peligro)
        }
    
    # Fallback (no debería ocurrir)
    return {
        "type": "Desconocido",
        "icon": "❓",
        "description": f"Consumo no catalogado: {irms:.2f}A",
        "color": "#95a5a6"
    }


# ======================================================================
# UMBRALES CONFIGURABLES POR SENSOR (Sin cambios)
# ======================================================================

def detect_device_type(irms: float, threshold: float) -> dict:
    """
    Clasifica el estado electrico de un ramal de 2 PCs.
    Retorna: {type, icon, description, color}
    """
    if irms >= threshold:
        if irms >= 15.0:
            return {
                "type": "Pico extremo del ramal",
                "icon": "ALRT",
                "description": f"Falla grave o pico electrico en el ramal ({irms:.2f}A)",
                "color": "#ff0000",
            }
        return {
            "type": "Sobrecarga del ramal",
            "icon": "ALRT",
            "description": f"El ramal ({irms:.2f}A) supera el umbral ({threshold:.1f}A)",
            "color": "#e74c3c",
        }

    if irms < 0.01:
        return {
            "type": "Sin carga",
            "icon": "OFF",
            "description": "Ramal sin consumo medible",
            "color": "#95a5a6",
        }
    if irms < BRANCH_RESIDUAL_MAX_CURRENT_A:
        return {
            "type": "Consumo residual",
            "icon": "STBY",
            "description": f"Consumo compatible con 2 PCs apagadas o en espera ({irms:.3f}A)",
            "color": "#3498db",
        }
    if irms < BRANCH_ONE_PC_MAX_CURRENT_A:
        return {
            "type": "1 PC encendida",
            "icon": "1PC",
            "description": f"Consumo compatible con 1 PC encendida en el ramal ({irms:.3f}A)",
            "color": "#27ae60",
        }
    if irms < BRANCH_ONE_PC_WORKLOAD_MAX_CURRENT_A:
        return {
            "type": "1 PC con programas",
            "icon": "1PC",
            "description": f"Consumo compatible con 1 PC ejecutando programas y la otra apagada ({irms:.3f}A)",
            "color": "#f39c12",
        }
    if irms < BRANCH_TWO_PC_MAX_CURRENT_A:
        return {
            "type": "2 PCs encendidas",
            "icon": "2PC",
            "description": f"Consumo compatible con ambas PCs usando programas ({irms:.3f}A)",
            "color": "#f39c12",
        }
    return {
        "type": "Ramal en uso alto",
        "icon": "HIGH",
        "description": f"Uso alto en ramal de 2 PCs ({irms:.3f}A)",
        "color": "#e67e22",
    }


def _parse_csv_env(name: str, default: str) -> list[str]:
    raw_value = os.getenv(name)
    source_value = raw_value if raw_value is not None else default
    values = [value.strip() for value in source_value.split(",") if value.strip()]
    if values:
        return values
    return [value.strip() for value in default.split(",") if value.strip()]


def _get_float_env(name: str, default: float) -> float:
    return _read_float_env(name, default)


DEFAULT_BRANCH_IDS = [f"C-{index:02d}" for index in range(1, 11)]
DEFAULT_BRANCH_ID_LIST = ",".join(DEFAULT_BRANCH_IDS)

LAB_ROOM_ID = os.getenv("LAB_ROOM_ID", "LAB-PC-01").strip() or "LAB-PC-01"
LAB_ROOM_NAME = os.getenv("LAB_ROOM_NAME", "Laboratorio de Computo").strip() or "Laboratorio de Computo"
SENSOR_IDS = _parse_csv_env("MONITORED_SENSOR_IDS", DEFAULT_BRANCH_ID_LIST)
ROOM_LABELS = {
    sensor_id: LAB_ROOM_NAME
    for sensor_id in SENSOR_IDS
}
SCHEDULE_STORE_PATH = os.getenv("SCHEDULE_STORE_PATH", "/config/schedules").rstrip("/")
OUT_OF_SCHEDULE_MIN_CURRENT_A = _get_float_env(
    "OUT_OF_SCHEDULE_MIN_CURRENT_A",
    BRANCH_RESIDUAL_MAX_CURRENT_A,
)
NO_CLASS_SCHEDULE_KINDS = {"no_class", "holiday", "suspension"}
USER_STORE_PATH = os.getenv("USER_STORE_PATH", "/app_users").rstrip("/")
TERMS_CONSENT_STORE_PATH = os.getenv("TERMS_CONSENT_STORE_PATH", "/app_consents").rstrip("/")
TERMS_VERSION = os.getenv("TERMS_VERSION", "2026-05-31")
ALERT_RECIPIENT_ROLES = {
    role.strip().lower()
    for role in os.getenv("ALERT_RECIPIENT_ROLES", "admin,auditor").split(",")
    if role.strip()
}
LOCAL_TIMEZONE = timezone(timedelta(hours=_get_float_env("APP_LOCAL_UTC_OFFSET_HOURS", -5.0)))

DEFAULT_THRESHOLDS = {
    sensor_id: {"corriente": 11.0, "potencia": 2420.0}
    for sensor_id in SENSOR_IDS
}

def get_sensor_threshold(sensor_id: str) -> dict:
    try:
        ref = db.reference(f'/config/thresholds/{sensor_id}')
        threshold = ref.get()
        if threshold:
            return threshold
        else:
            default = DEFAULT_THRESHOLDS.get(sensor_id, {"corriente": 11.0, "potencia": 2420.0})
            ref.set(default) 
            return default
    except Exception as e:
        print(f"Error al obtener umbral: {str(e)}")
        return DEFAULT_THRESHOLDS.get(sensor_id, {"corriente": 11.0, "potencia": 2420.0})

def update_sensor_threshold(sensor_id: str, corriente: float, potencia: float) -> bool:
    try:
        ref = db.reference(f'/config/thresholds/{sensor_id}')
        ref.set({
            "corriente": corriente,
            "potencia": potencia,
            "updated_at": datetime.now().isoformat()
        })
        return True
    except Exception as e:
        print(f"Error al actualizar umbral: {str(e)}")
        return False


def _schedule_path(room_id: str, schedule_id: Optional[str] = None) -> str:
    base_path = f"{SCHEDULE_STORE_PATH}/{room_id}"
    if schedule_id:
        return f"{base_path}/{schedule_id}"
    return base_path


def list_room_schedules(room_id: Optional[str] = None) -> List[Dict[str, Any]]:
    try:
        if room_id:
            data = db.reference(_schedule_path(room_id)).get()
            return _records_from_schedule_node(room_id, data)

        data = db.reference(SCHEDULE_STORE_PATH).get()
        schedules: List[Dict[str, Any]] = []
        if isinstance(data, dict):
            for current_room_id, room_data in data.items():
                schedules.extend(_records_from_schedule_node(str(current_room_id), room_data))
        return schedules
    except Exception as e:
        print(f"Error al listar horarios: {str(e)}")
        return []


def _records_from_schedule_node(room_id: str, data: Any) -> List[Dict[str, Any]]:
    if not isinstance(data, dict):
        return []

    records: List[Dict[str, Any]] = []
    for schedule_id, record in data.items():
        if isinstance(record, dict):
            item = dict(record)
            item.setdefault("id", str(schedule_id))
            item.setdefault("room_id", room_id)
            records.append(item)
    records.sort(
        key=lambda item: (
            str(item.get("day_of_week", "")),
            str(item.get("start_time", "")),
            0 if str(item.get("kind", "class")).lower() != "no_class" else 1,
            str(item.get("valid_from", "")),
            str(item.get("label", "")),
        )
    )
    return records


def save_room_schedule(room_id: str, schedule_id: str, schedule_record: Mapping[str, Any]) -> Dict[str, Any]:
    record = dict(schedule_record)
    record["id"] = schedule_id
    record["room_id"] = room_id
    db.reference(_schedule_path(room_id, schedule_id)).set(record)
    return record


def update_room_schedule(room_id: str, schedule_id: str, schedule_patch: Mapping[str, Any]) -> Dict[str, Any]:
    ref = db.reference(_schedule_path(room_id, schedule_id))
    current_record = ref.get()
    if not isinstance(current_record, dict):
        raise KeyError("Horario no encontrado")

    updated_record = dict(current_record)
    updated_record.update(dict(schedule_patch))
    updated_record["id"] = schedule_id
    updated_record["room_id"] = room_id
    ref.set(updated_record)
    return updated_record


def _schedule_kind(schedule: Mapping[str, Any]) -> str:
    kind = str(schedule.get("kind") or "class").strip().lower()
    if kind in NO_CLASS_SCHEDULE_KINDS:
        return "no_class"
    return "class"


def _schedule_matches_date(schedule: Mapping[str, Any], day_name: str, current_date: str) -> bool:
    if str(schedule.get("status", "activo")).lower() != "activo":
        return False
    if str(schedule.get("day_of_week", "")).lower() != day_name:
        return False

    valid_from = str(schedule.get("valid_from") or "")
    valid_to = str(schedule.get("valid_to") or "")
    if valid_from and current_date < valid_from:
        return False
    if valid_to and current_date > valid_to:
        return False
    return True


def _schedule_matches_clock(schedule: Mapping[str, Any], current_clock: str) -> bool:
    start_time = str(schedule.get("start_time") or "")
    end_time = str(schedule.get("end_time") or "")
    return bool(start_time and end_time and start_time <= current_clock <= end_time)


def get_schedule_context(room_id: str, when: Optional[datetime] = None) -> Dict[str, Any]:
    current_time = when or datetime.now()
    day_name = current_time.strftime("%A").lower()
    current_clock = current_time.strftime("%H:%M")
    current_date = current_time.date().isoformat()
    schedules = list_room_schedules(room_id)

    for schedule in schedules:
        if _schedule_kind(schedule) == "no_class" and _schedule_matches_date(schedule, day_name, current_date) and _schedule_matches_clock(schedule, current_clock):
            return {
                "is_scheduled_now": False,
                "blocked_by_no_class": True,
                "label": "Dia sin clase",
            }

    for schedule in schedules:
        if _schedule_kind(schedule) == "class" and _schedule_matches_date(schedule, day_name, current_date) and _schedule_matches_clock(schedule, current_clock):
            return {
                "is_scheduled_now": True,
                "blocked_by_no_class": False,
                "label": "En horario",
            }

    return {
        "is_scheduled_now": False,
        "blocked_by_no_class": False,
        "label": "Sin horario activo",
    }


def is_room_in_allowed_schedule(room_id: str, when: Optional[datetime] = None) -> bool:
    return bool(get_schedule_context(room_id, when).get("is_scheduled_now"))


def get_schedule_status(room_id: str, irms: float, when: Optional[datetime] = None) -> Dict[str, Any]:
    schedule_context = get_schedule_context(room_id, when)
    is_scheduled_now = bool(schedule_context["is_scheduled_now"])
    is_out_of_schedule = irms >= OUT_OF_SCHEDULE_MIN_CURRENT_A and not is_scheduled_now
    return {
        "is_scheduled_now": is_scheduled_now,
        "is_out_of_schedule": is_out_of_schedule,
        "blocked_by_no_class": bool(schedule_context["blocked_by_no_class"]),
        "min_current_a": OUT_OF_SCHEDULE_MIN_CURRENT_A,
        "label": "Fuera de horario" if is_out_of_schedule else str(schedule_context["label"]),
    }

# ======================================================================
# ¡FUNCIÓN DE LECTURA MODIFICADA!
# ======================================================================

def _now_pair() -> tuple[datetime, datetime]:
    now_utc = datetime.now(timezone.utc).replace(microsecond=0)
    return now_utc, now_utc.astimezone(LOCAL_TIMEZONE)


def _firebase_safe_key(value: str) -> bool:
    return bool(value) and not any(character in value for character in {".", "#", "$", "[", "]", "/"})


def _history_key(now_utc: datetime) -> str:
    return f"{now_utc.strftime('%Y%m%dT%H%M%SZ')}_{uuid.uuid4().hex[:8]}"


def _float_value(value: object, default: float) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _threshold_value(threshold: Mapping[str, Any], field_name: str, default: float) -> float:
    return _float_value(threshold.get(field_name), default)


def _estado_from_flags(is_overload: bool, is_out_of_schedule: bool) -> str:
    if is_overload:
        return "Sobrecarga"
    if is_out_of_schedule:
        return "Fuera de horario"
    return "Normal"


def record_iot_reading(
    sensor_id: str,
    irms: float,
    potencia: Optional[float] = None,
    *,
    voltage: float = 220.0,
    circuito: Optional[str] = None,
) -> Dict[str, Any]:
    sensor_id = sensor_id.strip()
    if sensor_id not in SENSOR_IDS:
        raise ValueError(f"Sensor no monitoreado: {sensor_id}")
    if not _firebase_safe_key(sensor_id):
        raise ValueError("El identificador del sensor contiene caracteres no permitidos")
    if irms < 0:
        raise ValueError("irms no puede ser negativo")
    if potencia is not None and potencia < 0:
        raise ValueError("potencia no puede ser negativa")
    if voltage <= 0:
        raise ValueError("voltage debe ser mayor que cero")

    now_utc, now_local = _now_pair()
    threshold = get_sensor_threshold(sensor_id)
    measured_power = round(potencia if potencia is not None else irms * voltage, 3)
    measured_current = round(irms, 3)
    current_threshold = _threshold_value(threshold, "corriente", 11.0)
    power_threshold = _threshold_value(threshold, "potencia", current_threshold * voltage)
    is_overload = measured_current >= current_threshold or measured_power >= power_threshold
    schedule_status = get_schedule_status(LAB_ROOM_ID, measured_current, now_local)
    is_out_of_schedule = bool(schedule_status["is_out_of_schedule"])
    estado = _estado_from_flags(is_overload, is_out_of_schedule)
    device_info = detect_device_type(measured_current, current_threshold)

    current_record = {
        "circuito": circuito or sensor_id,
        "estado": estado,
        "irms": measured_current,
        "potencia": measured_power,
        "timestamp": now_local.isoformat(),
        "timestamp_utc": now_utc.isoformat().replace("+00:00", "Z"),
        "updated_at": now_utc.isoformat(),
        "schedule_room_id": LAB_ROOM_ID,
    }
    history_record = {
        **current_record,
        "room_name": ROOM_LABELS.get(sensor_id, sensor_id),
        "is_overload": is_overload,
        "is_out_of_schedule": is_out_of_schedule,
        "schedule": schedule_status,
        "threshold": threshold,
    }
    history_key = _history_key(now_utc)

    db.reference(f"/current_data/{sensor_id}").set(current_record)
    db.reference(f"/history/{sensor_id}/{history_key}").set(history_record)

    return {
        "id": sensor_id,
        "room_name": ROOM_LABELS.get(sensor_id, sensor_id),
        "circuito": current_record["circuito"],
        "irms": measured_current,
        "potencia": measured_power,
        "is_overload": is_overload,
        "is_out_of_schedule": is_out_of_schedule,
        "timestamp": current_record["timestamp"],
        "timestamp_utc": current_record["timestamp_utc"],
        "schedule_room_id": current_record["schedule_room_id"],
        "estado": estado,
        "device": device_info,
        "threshold": threshold,
        "schedule": schedule_status,
        "history_key": history_key,
    }


def _user_has_current_terms_consent(username: str, uid: str, role: str) -> bool:
    if not _firebase_safe_key(username):
        return False
    try:
        records = db.reference(f"{TERMS_CONSENT_STORE_PATH}/{username}").get()
    except Exception as e:
        print(f"Error al leer consentimientos: {str(e)}")
        return False

    if not isinstance(records, dict):
        return False

    for record in records.values():
        if not isinstance(record, dict):
            continue
        if str(record.get("event_type", "terms_acceptance")) != "terms_acceptance":
            continue
        if str(record.get("terms_version", "")) != TERMS_VERSION:
            continue
        record_uid = str(record.get("uid") or "")
        record_role = str(record.get("role") or "").lower()
        if record_uid and uid and record_uid != uid:
            continue
        if record_role and record_role != role:
            continue
        return True
    return False


def _alert_contact_name(username: str, user_record: Mapping[str, Any]) -> str:
    for field_name in ("full_name", "display_name", "name", "username"):
        value = str(user_record.get(field_name) or "").strip()
        if value:
            return value
    return username


def get_alert_email_contacts(roles: Sequence[str] | None = None) -> List[Dict[str, str]]:
    allowed_roles = {role.lower() for role in (roles or ALERT_RECIPIENT_ROLES)}
    try:
        users = db.reference(USER_STORE_PATH).get()
    except Exception as e:
        print(f"Error al leer destinatarios de alerta: {str(e)}")
        return []

    if not isinstance(users, dict):
        return []

    contacts: dict[str, Dict[str, str]] = {}
    for username, user_record in users.items():
        if not isinstance(user_record, dict):
            continue

        role = str(user_record.get("role") or "").strip().lower()
        email = str(user_record.get("email") or "").strip()
        status = str(user_record.get("status") or "").strip().lower()
        uid = str(user_record.get("uid") or "").strip()
        is_disabled = bool(user_record.get("disabled", False))

        if role not in allowed_roles or status != "activo" or is_disabled or not email:
            continue
        if "@" not in email:
            continue
        if not _user_has_current_terms_consent(str(username), uid, role):
            continue

        contacts[email.lower()] = {
            "email": email,
            "name": _alert_contact_name(str(username), user_record),
            "username": str(username),
            "role": role,
        }

    return sorted(contacts.values(), key=lambda contact: contact["email"].lower())


def get_alert_email_recipients(roles: Sequence[str] | None = None) -> List[str]:
    return [contact["email"] for contact in get_alert_email_contacts(roles)]


def get_current_data() -> dict:
    """
    Obtiene los datos actuales con detección de dispositivos MEJORADA.
    """
    try:
        ref = db.reference('/current_data')
        data = ref.get()
        
        sensors_data = []
        any_connected = False
        total_consumption = 0
        
        if data:
            for sensor_id in SENSOR_IDS:
                # 1. Obtener umbral específico del sensor
                threshold = get_sensor_threshold(sensor_id)
                
                if sensor_id in data:
                    sensor_info = data[sensor_id]
                    irms = float(sensor_info.get('irms', 0.0))
                    potencia = float(sensor_info.get('potencia', 0.0))
                    
                    # 2. Determinar si hay sobrecarga
                    current_threshold = _threshold_value(threshold, "corriente", 11.0)
                    power_threshold = _threshold_value(threshold, "potencia", current_threshold * 220.0)
                    is_overload = irms >= current_threshold or potencia >= power_threshold
                    
                    # 3. ¡LÓGICA MEJORADA!
                    #    Pasamos el 'irms' Y el 'threshold' a la función
                    device_info = detect_device_type(irms, current_threshold)
                    schedule_status = get_schedule_status(LAB_ROOM_ID, irms)
                    
                    sensors_data.append({
                        "id": sensor_id,
                        "room_name": ROOM_LABELS.get(sensor_id, sensor_id),
                        "circuito": sensor_info.get("circuito", sensor_id),
                        "irms": irms,
                        "potencia": potencia,
                        "is_overload": is_overload,
                        "is_out_of_schedule": schedule_status["is_out_of_schedule"],
                        "timestamp": sensor_info.get('timestamp', ''),
                        "schedule_room_id": sensor_info.get("schedule_room_id", LAB_ROOM_ID),
                        "device": device_info,       # Info del dispositivo (ahora es más inteligente)
                        "threshold": threshold,      # Info del umbral
                        "schedule": schedule_status
                    })
                    
                    total_consumption += potencia
                    any_connected = True
                else:
                    # Sensor sin datos
                    sensors_data.append({
                        "id": sensor_id, "room_name": ROOM_LABELS.get(sensor_id, sensor_id),
                        "circuito": sensor_id,
                        "irms": 0.0, "potencia": 0.0, "is_overload": False,
                        "is_out_of_schedule": False,
                        "timestamp": "", 
                        "device": detect_device_type(0.0, threshold["corriente"]), # "Sin carga"
                        "threshold": threshold,
                        "schedule_room_id": LAB_ROOM_ID,
                        "schedule": get_schedule_status(LAB_ROOM_ID, 0.0)
                    })
            
            return {
                "sensors": sensors_data, "connected": any_connected,
                "message": "Sistema activo" if any_connected else "Sin dispositivos conectados",
                "timestamp": datetime.now().isoformat(), "total_consumption": total_consumption
            }
        else:
            # No hay datos en Firebase
            return {
                "sensors": [
                    {
                        "id": sid, "room_name": ROOM_LABELS.get(sid, sid),
                        "circuito": sid,
                        "irms": 0.0, "potencia": 0.0, "is_overload": False,
                        "is_out_of_schedule": False,
                        "timestamp": "", "device": detect_device_type(0.0, get_sensor_threshold(sid)["corriente"]),
                        "threshold": get_sensor_threshold(sid),
                        "schedule_room_id": LAB_ROOM_ID,
                        "schedule": get_schedule_status(LAB_ROOM_ID, 0.0)
                    } for sid in SENSOR_IDS
                ],
                "connected": False, "message": "Sin datos disponibles",
                "timestamp": datetime.now().isoformat(), "total_consumption": 0
            }
            
    except Exception as e:
        print(f"Error al obtener datos de Firebase: {str(e)}")
        # Devuelve una estructura de error que el frontend pueda manejar
        return {
            "sensors": [
                {
                    "id": sid, "room_name": ROOM_LABELS.get(sid, sid),
                    "circuito": sid,
                    "irms": 0.0, "potencia": 0.0, "is_overload": False,
                    "is_out_of_schedule": False,
                    "timestamp": "", "device": detect_device_type(0.0, get_sensor_threshold(sid)["corriente"]),
                    "threshold": get_sensor_threshold(sid),
                    "schedule_room_id": LAB_ROOM_ID,
                    "schedule": get_schedule_status(LAB_ROOM_ID, 0.0)
                } for sid in SENSOR_IDS
            ],
            "connected": False, "message": f"Error de conexión: {str(e)}",
            "timestamp": datetime.now().isoformat(), "total_consumption": 0
        }

# ======================================================================
# ¡FUNCIONES DE HISTORIAL Y ALERTAS MODIFICADAS!
# ======================================================================

def _record_timestamp(record_key: str, record: Mapping[str, Any]) -> str:
    timestamp = str(record.get("timestamp") or "")
    timestamp_utc = str(record.get("timestamp_utc") or "")
    if len(timestamp) >= 16:
        return timestamp
    if len(timestamp_utc) >= 16:
        return timestamp_utc
    return record_key


def _parse_datetime_utc(value: object, *, assume_local: bool, end_of_day: bool = False) -> Optional[datetime]:
    raw_value = unquote(str(value or "").strip())
    if not raw_value:
        return None
    if len(raw_value) == 10:
        raw_value = f"{raw_value}T{'23:59:59.999999' if end_of_day else '00:00:00'}"
    try:
        parsed = datetime.fromisoformat(raw_value.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=LOCAL_TIMEZONE if assume_local else timezone.utc)
    return parsed.astimezone(timezone.utc)


def _record_datetime_utc(record_key: str, record: Mapping[str, Any]) -> Optional[datetime]:
    timestamp_utc = _parse_datetime_utc(record.get("timestamp_utc"), assume_local=False)
    if timestamp_utc is not None:
        return timestamp_utc
    timestamp = _parse_datetime_utc(record.get("timestamp"), assume_local=True)
    if timestamp is not None:
        return timestamp
    return _parse_datetime_utc(record_key, assume_local=True)


def _within_date_range(record_key: str, record: Mapping[str, Any], start_date: str = None, end_date: str = None) -> bool:
    record_datetime = _record_datetime_utc(record_key, record)
    start_datetime = _parse_datetime_utc(start_date, assume_local=True)
    end_datetime = _parse_datetime_utc(end_date, assume_local=True, end_of_day=True)
    if (start_datetime or end_datetime) and record_datetime is None:
        return False
    if start_datetime and record_datetime and record_datetime < start_datetime:
        return False
    if end_datetime and record_datetime and record_datetime > end_datetime:
        return False
    return True


def _is_reportable_history_record(record: Mapping[str, Any]) -> bool:
    estado = str(record.get("estado") or "Normal")
    return estado in {"Sobrecarga", "Fuera de horario"} or bool(record.get("is_overload")) or bool(record.get("is_out_of_schedule"))


def get_history_data(
    sensor_id: str,
    limit: int = 20,
    start_date: str = None,
    end_date: str = None,
    reportable_only: bool = False,
) -> List[Dict]:
    """
    Obtiene el historial con filtros de fecha (HU-010)
    """
    try:
        ref = db.reference(f'/history/{sensor_id}')
        threshold = get_sensor_threshold(sensor_id)["corriente"] # Obtener umbral
        
        data = ref.order_by_key().get() if start_date or end_date else ref.order_by_key().limit_to_last(limit).get()
        
        if data:
            history = []
            for key, value in data.items():
                if not isinstance(value, dict):
                    continue
                if reportable_only and not _is_reportable_history_record(value):
                    continue
                timestamp = _record_timestamp(key, value)
                if not _within_date_range(key, value, start_date, end_date):
                    continue
                irms = float(value.get('irms', 0.0))
                # Usamos la nueva lógica de detección aquí también
                device_info = detect_device_type(irms, threshold) 
                
                history.append({
                    "id": key,
                    "timestamp": timestamp,
                    "timestamp_utc": value.get("timestamp_utc", ""),
                    "irms": irms,
                    "potencia": float(value.get('potencia', 0.0)),
                    "estado": value.get('estado', 'Normal'), # El estado sigue siendo el mismo
                    "device": device_info,
                    "_sort_at": (_record_datetime_utc(key, value) or datetime.min.replace(tzinfo=timezone.utc)).isoformat(),
                })
            history.sort(key=lambda record: record["_sort_at"], reverse=True)
            for record in history:
                record.pop("_sort_at", None)
            return history
        return []
    except Exception as e:
        print(f"Error al obtener historial: {str(e)}")
        return []

def get_alert_history(start_date: str = None, end_date: str = None) -> List[Dict]:
    """
    Obtiene un historial de ÚNICAMENTE los eventos de sobrecarga.
    """
    try:
        all_alerts = []
        alert_states = {
            "Sobrecarga": "overload",
            "Fuera de horario": "out_of_schedule_consumption",
        }
        for sensor_id in SENSOR_IDS:
            ref = db.reference(f'/history/{sensor_id}')
            threshold = get_sensor_threshold(sensor_id) # Obtener umbral como dict
            
            data = ref.order_by_key().get()
            
            if data:
                for key, value in data.items():
                    if not isinstance(value, dict):
                        continue
                    estado = str(value.get('estado', 'Normal'))
                    if estado not in alert_states:
                        continue
                    timestamp = _record_timestamp(key, value)
                    if not _within_date_range(key, value, start_date, end_date):
                        continue

                    irms = float(value.get('irms', 0.0))
                    # Usamos la nueva lógica de detección aquí también
                    device_info = detect_device_type(irms, threshold["corriente"])
                    
                    all_alerts.append({
                        "id": key,
                        "sensor_id": sensor_id,
                        "timestamp": timestamp,
                        "timestamp_utc": value.get("timestamp_utc", ""),
                        "alert_type": alert_states[estado],
                        "irms": irms,
                        "potencia": float(value.get('potencia', 0.0)),
                        "estado": estado,
                        "device": device_info,
                        "threshold": threshold,
                        "_sort_at": (_record_datetime_utc(key, value) or datetime.min.replace(tzinfo=timezone.utc)).isoformat(),
                    })

        all_alerts.sort(key=lambda x: x["_sort_at"], reverse=True)
        for alert in all_alerts:
            alert.pop("_sort_at", None)
        return all_alerts
        
    except Exception as e:
        print(f"Error al obtener historial de alertas: {str(e)}")
        return []

# ======================================================================
# FUNCIONES DE EXPORTACIÓN (Sin cambios)
# ======================================================================

def export_history_csv(sensor_id: str = None, start_date: str = None, end_date: str = None, reportable_only: bool = False) -> str:
    try:
        if sensor_id:
            sensors = [sensor_id]
        else:
            sensors = SENSOR_IDS
        
        csv_data = "Sensor ID,Fecha/Hora,Corriente (A),Potencia (W),Dispositivo,Estado\n"
        
        for sid in sensors:
            history = get_history_data(sid, limit=1000, start_date=start_date, end_date=end_date, reportable_only=reportable_only)
            for record in history:
                csv_data += f"{sid},{record['timestamp']},{record['irms']:.3f},{record['potencia']:.2f},{record['device']['type']},{record['estado']}\n"
        
        return csv_data
    except Exception as e:
        print(f"Error al exportar CSV: {str(e)}")
        return ""

def check_connection() -> bool:
    try:
        ref = db.reference('/current_data')
        ref.get()
        return True
    except Exception as e:
        print(f"Error al verificar conexión: {str(e)}")
        return False
    
def export_history_excel(sensor_id: str = None, start_date: str = None, end_date: str = None, reportable_only: bool = False) -> bytes:
    try:
        if sensor_id:
            sensors = [sensor_id]
        else:
            sensors = SENSOR_IDS
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial SafyraShield"
        
        header_font = Font(bold=True, color="FFFFFF", name="Inter")
        header_fill = PatternFill(start_color="0A0E27", end_color="0A0E27", fill_type="solid")
        cell_font = Font(name="Inter")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        headers = ["Sensor ID", "Fecha/Hora (ISO)", "Corriente (A)", "Potencia (W)", "Dispositivo Detectado", "Estado"]
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
            ws.row_dimensions[1].height = 25

        row_idx = 2
        for sid in sensors:
            history = get_history_data(sid, limit=1000, start_date=start_date, end_date=end_date, reportable_only=reportable_only)
            for record in history:
                row_data = [
                    sid,
                    record['timestamp'],
                    record['irms'],
                    record['potencia'],
                    record['device']['type'],
                    record['estado']
                ]
                ws.append(row_data)
                
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.border = thin_border
                    cell.font = cell_font
                    if col_num in [1, 5, 6]: cell.alignment = center_align
                    else: cell.alignment = left_align
                    if col_num == 3: cell.number_format = '0.000 "A"'
                    if col_num == 4: cell.number_format = '0.00 "W"'
                
                ws.row_dimensions[row_idx].height = 20
                row_idx += 1
        
        ws.column_dimensions[get_column_letter(1)].width = 15
        ws.column_dimensions[get_column_letter(2)].width = 28
        ws.column_dimensions[get_column_letter(3)].width = 15
        ws.column_dimensions[get_column_letter(4)].width = 15
        ws.column_dimensions[get_column_letter(5)].width = 25
        ws.column_dimensions[get_column_letter(6)].width = 12
        
        with io.BytesIO() as buffer:
            wb.save(buffer)
            return buffer.getvalue()

    except Exception as e:
        print(f"Error al exportar Excel: {str(e)}")
        return b""
